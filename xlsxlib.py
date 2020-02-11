import os, sys
import logging

logger = logging.getLogger(__package__)

import openpyxl
from openpyxl.styles import Font

def mangle_sheet_name(sheet_name):
    "Sheet names cannot be longer than 31 chars"
    return sheet_name[:31]

def cells_in_range(range):
    for row in range:
        for cell in row:
            yield cell

def xlsx(data_iterator, spreadsheet_filepath):
    """xlsx - put a dataset to an xlsx spreadsheet

    Parameters:
        an iterator which will supply [(Sheet Name, [Column Names / Types], [Rows]), ...]
        spreadsheet_filepath - full path to a spreadsheet
    """
    ROWSET_SIZE = 1000

    wb = openpyxl.Workbook()
    for sheet in list(wb.worksheets):
        wb.remove_sheet(sheet)

    for n_sheet, (sheet_name, headers, rowset) in enumerate(data_iterator):
        #
        # Create a new sheet to hold the rowset. Use the sheet name supplied,
        # adjusted as necessary to meet the constraints of Excel sheet names,
        # or create a default one if none is supplied.
        #
        if sheet_name:
            sheet_name = mangle_sheet_name(sheet_name)
        else:
            sheet_name = u"Sheet %d" % n_sheet
        ws = wb.create_sheet(title=sheet_name)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        yield "%s... " % sheet_name

        #
        # Write a single row containing the headers
        # Make the headers bold and freeze panes below that row
        #
        header_names = [name for name, type in headers]        

        ws.append(header_names)
        ws.freeze_panes = "A2"

        bold = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = cell.font = bold    
        yield "%s headers..." % sheet_name

        #
        # Append each row to the bottom of the sheet
        #
        n_row = 0
        for n_row, row in enumerate(rowset):
            if (1 + n_row) % ROWSET_SIZE == 0:
                yield "%s row %d" % (sheet_name, n_row)
            ws.append(list(row))
        yield "%s %d rows" % (sheet_name, n_row)

        #
        # Get the max characters in each column. Then set the width to 
        # that. Width is exactly the width of a monospace font (if not
        # changing other styles at least). Even if you use a variable width 
        # font it is a decent estimation. This will not work with formulas.
        #
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column].width = adjusted_width


    wb.save(filename=spreadsheet_filepath)
