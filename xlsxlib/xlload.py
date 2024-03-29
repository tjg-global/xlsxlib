#!python3
"""xlload - load an Excel file into a database table
"""
import os, sys
import argparse
import csv
import datetime
import getpass
import itertools
import logging
import re
import tempfile
import time

import openpyxl

from . import connections

NAME, _ = os.path.splitext(os.path.basename(__file__))
UNCONVERTED = object()
TRUNCATE = True

def as_code(name):
    return "_".join(name.lower().split())

_converters = {
    datetime.datetime : lambda v: v.strftime("%d %b %Y"),
    datetime.date : lambda v: v.strftime("%d %b %Y"),
    type(None) : lambda v: "",
    str: lambda v: v.encode("utf-8"),
}

def sheet_metadata(sheet):
    """Return useful metadata from the Excel sheet

    Read the first line to retrieve header names
    Read the next (non-blank) line to retrieve datatype info
    Determine the number of rows

    Return [headers], [types], n_rows
    """
    irows = sheet.iter_rows()
    header_values = (c.value for c in next(irows))
    headers = list(itertools.takewhile(lambda v: v is not None, header_values))
    types = [type("" if c.value is None else c.value) for (c, _) in zip(next(irows), headers)]
    remaining_rows = list(row for row in irows if any(c.value for c in row))
    return headers, types, 1 + len(remaining_rows)

def sheet_as_rows(sheet):
    for n_row, row in enumerate(sheet):
        values = [cell.value for cell in row]
        if n_row == 0:
            header = list(itertools.takewhile(lambda v: v is not None, values))
        truncated_values = values[:len(header)]
        if any(v is not None for v in truncated_values):
            yield [_converters.get(type(value), str)(value) for value in truncated_values]
        else:
            break

def sheet_from_xlsx(xlsx_filepath, sheet_name):
    """Quick hack to open openpyxl to convert .xls[x] files to .csv
    """
    workbook = openpyxl.load_workbook(
        xlsx_filepath,
        data_only=True, read_only=True
    )
    #
    # If a sheet name is specified use that, otherwise use the first sheet
    # in the workbook
    #
    if sheet_name:
        return workbook[sheet_name]
    else:
        for sheet in workbook:
            return sheet

    #
    # Assume the first row is representative.
    # If the first blank cell is less than the supposed max column, assume
    # that max column is flawed and bring it down. We could do the same for rows,
    # but we're already filtering blank rows out so it's not so much of a problem.
    #
    # NB use the sheet iterator (which yields rows) rather than the .rows
    # iterator as the former restarts while the latter doesn't
    #
    for row in sheet:
        break
    for n_column, cell in enumerate(row):
        if cell.value is None:
            sheet.max_column = min(sheet.max_column, n_column)
            break
    csv_filepath = tempfile.mktemp(".csv")
    with open(csv_filepath, "wb") as f:
        csv.writer(f).writerows(sheet_as_rows(sheet))
    return csv_filepath

class DictReader(csv.DictReader):
    """Our CSV files can have spaces around their field headers. Since these
    won't match with the corresponding database column names, universally strip
    the fields used by the DictReader.

    Also allow fields to mapped to other names so the .csv header can carry different
    names for legacy or readability reasons compared to the database tables.
    """
    def __init__(self, f, mappings, database_columns, *args, **kwargs):
        csv.DictReader.__init__(self, f, *args, **kwargs)
        fieldnames = [f.strip() for f in self.fieldnames if f.strip()]
        #
        # Look first for an explicit fieldname mapping
        # Then allow for an implicit mapping between "Field Name" and field_name
        #
        self._fieldnames = []
        for f in fieldnames:
            if f in mappings:
                self._fieldnames.append(mappings[f])
            elif as_code(f) in database_columns:
                self._fieldnames.append(as_code(f))
            else:
                self._fieldnames.append(f)

def collapse_whitespace(value):
    return re.sub(r"\s{2,}", " ", value)

def convert_to_datetime(value):
    #
    # Attempt a series of likely steps to convert something to
    # a datetime value.
    #
    # 1) If it's a blank or the word "NULL" (!), treat it as a NULL
    # 2) Try a few common date formats in turn
    # 3) If it's a number assume it's Excel's days-since-19000101 algorithm
    #
    if not value.strip():
        return None
    if value.strip().upper() in ("NULL", "?"):
        return None

    #
    # Try common date formats
    #
    for possible in "%d/%m/%Y", "%d %b %Y", "%Y%m%d", "%d/%m/%y":
        try:
            return datetime.datetime.strptime(value, possible)
        except ValueError:
            continue

    #
    # Try days-since-1st Jan 1900
    #
    try:
        n_days_since_19000101 = float(value)
        d = datetime.datetime(1900, 1, 1) + datetime.timedelta(days=n_days_since_19000101)
        #
        # Round up or down according to the number of microseconds
        #
        d2 = datetime.datetime(
            year=d.year, month=d.month, day=d.day,
            hour=d.hour, minute=d.minute
        )
        d3 = d2 + datetime.timedelta(seconds=int(d.second + (d.microsecond / 1000000.0) + 0.5))
        return d3
    except ValueError:
        pass

    raise ValueError("Unable to convert %s to a datetime value" % value)

def convert_to_time(value):
    dt = convert_to_datetime(value)
    if dt is None:
        return dt
    else:
        return dt.time()

def convert_to_number(value):
    value = str(value).strip()
    if value in ("-", ""):
        return None
    else:
        return value.replace(",", "")

def convert_to_float(value):
    number = convert_to_number(value)
    if number is None:
        return None
    else:
        return float(number)

def convert_to_int(value):
    number = convert_to_number(value)
    if number is None:
        return None
    else:
        return int(number)

bit_values = {
    "yes" : 1, "no" : 0,
    "y" : 1, "n" : 0,
    "true" : 1, "false" : 0,
    "" : 0
}
def convert_to_bit(value):
    return bool(bit_values.get(value.lower(), convert_to_number(value)))

def convert_to_unicode(value):
    #
    # The value from a CSV will be mbcs-encoded bytes
    #
    return unicode(value, "mbcs")

type_mappings = {
  "int" : convert_to_int,
  "tinyint" : convert_to_int,
  "bigint" : convert_to_int,
  "smallint" : convert_to_int,
  "bit" : convert_to_bit,
  "money" : convert_to_float,
  "numeric" : convert_to_float,
  "decimal" : convert_to_float,
  "smallmoney" : convert_to_float,
  "float" : convert_to_float,
  "real" : convert_to_float,
  "date" : convert_to_datetime,
  "datetime" : convert_to_datetime,
  "smalldatetime" : str,
  "char" : str,
  "varchar" : str,
  "text" : str,
  "nchar" : convert_to_unicode,
  "nvarchar" : convert_to_unicode,
  "ntext" : convert_to_unicode,
  "binary" : bytes,
  "varbinary" : bytes,
  "image" : bytes,
  "timestamp" : str,
  "uniqueidentifier" : str,
  "time" : convert_to_time
}

def fqon(db, object_name):
    return db.execute(
    """SELECT
        sch.name,
        obj.name
      FROM
        sys.objects AS obj
      JOIN sys.schemas AS sch ON
        sch.schema_id = obj.schema_id
      WHERE
        obj.object_id = OBJECT_ID(?)
    ;""",
    [object_name]).fetchone() or (None, None)

def table_columns(db, schema, table):
    q = db.cursor()
    try:
        for column in q.columns(table=table, schema=schema):
            name = column[3]
            datatype = column[5].lower()
            length = column[6] if datatype.endswith("char") else None
            is_nullable = column[17] == "YES"
            has_default = column[12] is not None or datatype.endswith("identity")

            yield name, (datatype, length, is_nullable, has_default)
    finally:
        q.close()

def validate_columns(database_columns, csv_columns, ignore_csv, ignore_database):
    """At the header level warn of trouble ahead: fields which are supplied
    for which there is no corresponding column in the table; and database
    columns for which there is no field in the CSV.
    """
    for column in csv_columns:
        if column not in database_columns:
            if column not in ignore_csv:
                yield "WARNING", "Column %s is in the CSV file but not in the database. It will not be loaded" % column
        else:
            datatype, length, is_nullable, has_default = database_columns[column]
            yield "DEBUG", "Column %s is of type %s" % (column, datatype)
    for column in set(database_columns) - set(csv_columns):
        if column not in ignore_database:
            yield "WARNING", "Column %s is in the database but missing from the CSV file" % column

def validate_row(row, converted_row, database_columns):
    """Validate the data in one CSV row against the corresponding database
    metadata. Only warnings are issued as it's not possible reliably to
    reproduce SQL Server's data parsing algorithm. The following checks are made:

    * That no data is supplied (ie the field isn't present) but the database column is NOT NULL
    * That the data in the field is unlikely to convert successfully to the datatype of the corresponding column
    * That the data in the field will likely overflow the size of the database column

    The INSERT will go ahead, regardless of any warnings issued here. If it fails,
    however, the data here should help diagnose the problem.
    """
    for column, (datatype, length, is_nullable, has_default) in database_columns.items():
        value = row.get(column)
        if value is not None:
            #
            # value is either missing (None) or a string so we can always strip
            #
            value = collapse_whitespace(value.strip())

            conversion_factory = type_mappings.get(datatype, str)
            try:
                converted_row[column] = conversion_factory(value)
            except (ValueError, TypeError):
                yield "WARNING", "Column %s has value %r which may not convert to %s" % (column, value, datatype)
                converted_row[column] = UNCONVERTED
            else:
                if length and len(value) > length:
                    yield "WARNING", "Column %s has value %r which may be too long for length %d" % (column, value, length)

        #
        # The value can be None as the result of a conversion (eg datetime)
        #
        if value is None:
            if not is_nullable and not has_default:
                yield "WARNING", "Column %s has no value but the database will not accept NULLs and has no default" % (column)

def _preprocess(iterator):
    """Apply any preprocessing steps immediately before processing the lines

    1) Replace any embedded lf characters with spaces. This comes from the Broadsign
       load which has embedded lf inside non-quoted fields. This prevents the
       standard Python file iterator from detecting the lines correctly, and
       the csv module will later raise an exception when it detects the
       situation.
    """
    for line in iterator:
        yield line.strip().replace("\n", " ")

def load_xlsx(db, logger, params):
    #~ schema, table = fqon(db, params.tablename)
    #~ if table is None:
        #~ logger.error("No such table: %s" % params.tablename)
        #~ raise RuntimeError
    #~ else:
        #~ fq_tablename = "[%s].[%s]" % (schema, table)
        #~ logger.debug("Writing into %s" % fq_tablename)

    #~ if params.truncate_first:
        #~ logger.debug("Truncate %s" % fq_tablename)
        #~ db.execute("TRUNCATE TABLE %s" % fq_tablename)
    #~ else:
        #~ logger.debug("Not truncating")

    #~ mappings = params.mappings or {}
    #~ if mappings:
        #~ for k, v in mappings.items():
            #~ logger.debug("Mapping %s to %s" % (k, v))

    #~ ignore_csv = params.ignore_csv or []
    #~ if ignore_csv:
        #~ logger.debug("Ignoring csv columns: %s" % (", ".join(ignore_csv)))

    #~ ignore_database = params.ignore_database or []
    #~ if ignore_database:
        #~ logger.debug("Ignoring database columns: %s" % (", ".join(ignore_database)))

    filepath = params.filepath
    if not os.path.exists(filepath):
        raise RuntimeError("Cannot find %s" % filepath)

    sheet = sheet_from_xlsx(params.filepath, params.sheet_name)
    headers, types, n_rows = sheet_metadata(sheet)
    print("Headers:", headers)
    print("Types:", types)
    print("Rows:", n_rows)
    return

    #
    # To give a "n of n loaded" message, determine how many lines are in the file
    #
    n_rows_in_file = 0

    database_columns = dict(table_columns(db, schema, table))
    #
    # Ignore any trailing columns, usually the artefact of an XLS -> CSV conversion
    #
    reader = DictReader(_preprocess(iterator), mappings, database_columns, delimiter=params.delimiter)
    #
    # At this stage we know the database columns and the columns in the CSV
    # file so we can validate to warn about missing / extra columns
    #
    for action, message in validate_columns(database_columns, reader.fieldnames, ignore_csv, ignore_database):
        logger.log(action, message)

    n_row = 0
    n_rows_loaded = 0
    n_rows_failed = 0
    q = db.cursor()
    q.fast_executemany = True
    try:
        #
        # For each row in the CSV file, perform some sanity checks on the values,
        # given the type of the corresponding database column, and then insert
        # the row. Carry on in the face of any database errors and log the total
        # number of inserted rows at the end
        #
        converted_rows = []
        for n, row in enumerate(reader):
            n_row = n + 1 ## make the counter human-readable
            converted_row = dict(row)
            for action, message in validate_row(row, converted_row, database_columns):
                logger.log(action, "Row %d: %s" % (n_row, message))

            #
            # The heart of the action: generate an INSERT statement which
            # uses all the columns which the CSV file supplies. If it succeeds,
            # add one notch to the successful load tally. If not, log an exception.
            # In either case, carry on.
            #
            common_fields = [f for f in converted_row if f in database_columns]
            if not common_fields:
                logger.error("None of the field in the CSV match fields in the database table")
                raise RuntimeError("No matching fields")

            insert_sql = "INSERT INTO %s (%s) VALUES (%s)" % (
                fq_tablename,
                ", ".join("[%s]" % name for name in common_fields),
                ", ".join("?" for _ in common_fields)
            )
            if UNCONVERTED in converted_row.values():
                logger.error("Unable to convert some values; row %d not loaded" % n_row)
                n_rows_failed += 1
            else:
                converted_rows.append([converted_row[f] for f in common_fields])

            is_batch_complete = (n_row % params.log_every_n_rows) == 0
            if is_batch_complete:
                try:
                    q.executemany(insert_sql, converted_rows)
                except KeyboardInterrupt:
                    raise
                except:
                    logger.exception("Rows not loaded" % (n_row, n_rows_in_file))
                    n_rows_failed += len(converted_rows)
                else:
                    n_rows_loaded += len(converted_rows)

                logger.debug("Row %d / %d loaded" % (n_rows_loaded, n_rows_in_file))
                converted_rows = []

            if n_rows_failed > 0 and params.fail_fast:
                break

        if converted_rows:
            try:
                q.executemany(insert_sql, converted_rows)
            except KeyboardInterrupt:
                raise
            except:
                logger.exception("Rows not loaded" % (n_row, n_rows_in_file))
                n_rows_failed += len(converted_rows)
            else:
                n_rows_loaded += len(converted_rows)

    finally:
        q.close()
        if closer:
            closer()

    logger.info("Loaded %d / %d rows" % (n_rows_loaded, n_rows_in_file))
    logger.debug("Loaded %d / %d rows into %s from %s" % (n_rows_loaded, n_rows_in_file, fq_tablename, filepath))
    if n_rows_failed:
        logger.error("%d rows could not be loaded" % n_rows_failed)

def main(db, params, logger):
    #
    # If no tablename is supplied, take the part of the CSV basename before the extension.
    # ie c:/temp/some.csv -> "some"
    #    c:/temp/datasources.some-other.csv -> "datasources.some-other"
    #
    if params.tablename is None:
        params.tablename, _ = os.path.splitext(os.path.basename(params.filepath))
    try:
        return load_xlsx(
            db, logger, params
        )
    except KeyboardInterrupt:
        logger.warn("KeyboardInterrupt: closing gracefully")
    except:
        #
        # Don't re-raise the error as that will result in a -1 return from the executable
        # which causes SSIS to produce a modal message box(!) saying that the return code
        # was -1
        #
        logger.exception("Some problem in the main process")

def get_db(args):
    driver, server_name, database_name, username, password = connections.parse_dburi_ex(args.dburi)
    if driver == "mssql":
        return connections.mssql(server_name, database_name, username, password)
    elif driver == "snowflake":
        return connections.snowflake(server_name, database_name, username, password)

class Params(object): pass

_delimiter_map = {
    "\\t" : "\t"
}

def get_params(args, logger):
    """Determine parameters from the command line, ini file or SSIS config

    The result is a params namespace object. Command line arguments always
    override; otherwise the only conflict to arise will be the database
    connection which has already been determined by this point in order
    to get the logger working.
    """
    params = Params()

    params.job_id = "%s-%s" % (getpass.getuser(), time.strftime("%Y%m%d-%H%M%S"))

    filepath = args.filepath
    if not filepath:
        raise RuntimeError("The path to a .xlsx file must be given")
    if "!" in filepath:
        filepath, _, params.sheet_name = filepath.partition("!")
    else:
        params.sheet_name = None
    params.filepath = filepath

    params.tablename = args.tablename

    if args.truncate is not None:
        truncate = args.truncate
    else:
        truncate = TRUNCATE
    params.truncate_first = bool(truncate)

    params.fields = []
    params.log_every_n_rows = args.log_every_n_rows
    if params.log_every_n_rows is None:
        params.log_every_n_rows = 1
    params.log_every_n_rows = int(params.log_every_n_rows)
    params.fail_fast = args.fail_fast or False

    return params

def command_line():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dburi", help="(for testing) mssql://[[user:]password@]server[/db]", required=True)
    parser.add_argument("--filepath", help="Full path to the xlsx file", required=True)
    parser.add_argument("--tablename", help="schema.tablename (default: same as the filename)")
    parser.add_argument('--truncate', dest='truncate', action='store_true')
    parser.add_argument('--no-truncate', dest='truncate', action='store_false')
    parser.add_argument('--log-every-n-rows', type=int, dest='log_every_n_rows')
    parser.add_argument('--fail-fast', help="Fail the load on the first error", action="store_true", dest="fail_fast")
    parser.set_defaults(truncate=None)
    args = parser.parse_args()

    logger = logging.getLogger(NAME)
    db = get_db(args)
    db.autocommit = True
    try:
        #
        # Try to use the job id from command line but fall back to "load_xlsx" so
        # at least it goes somewhere.
        #
        params = get_params(args, logger)
        main(db, params, logger)
    finally:
        db.close()

if __name__ == '__main__':
    command_line()
