import os, sys
from datetime import datetime
import logging

logger = logging.getLogger(__package__)

import openpyxl
from openpyxl.styles import Font, PatternFill

BOLD = Font(bold=True)
YELLOW = PatternFill("solid", fgColor="ffffff80")
INVALID_CHARS = str.maketrans("", "", "".join(chr(i) for i in range(32) if i not in (9, 10, 13)))

def mangle_sheet_name(sheet_name):
    "Sheet names cannot be longer than 31 chars"
    return sheet_name[:31]

def cells_in_range(range):
    for row in range:
        for cell in row:
            yield cell

def munged(row):
    """Translate certain datatypes which Excel won't handle
    """
    row = [(cell.translate(INVALID_CHARS) if isinstance(cell, str) else cell) for cell in row]
    row = [(cell.replace(tzinfo=None) if isinstance(cell, datetime) else cell) for cell in row]
    return row

def xlsx(data_iterator, spreadsheet_filepath):
    """xlsx - put a dataset to an xlsx spreadsheet

    Parameters:
        an iterator which will supply [(Sheet Name, [Column Names / Types], [Rows]), ...]
        spreadsheet_filepath - full path to a spreadsheet
    """
    ROWSET_SIZE = 1000

    wb = openpyxl.Workbook()
    for sheet in list(wb.worksheets):
        wb.remove_sheet(sheet)

    for n_sheet, (sheet_name, headers, rowset) in enumerate(data_iterator):
        #
        # Create a new sheet to hold the rowset. Use the sheet name supplied,
        # adjusted as necessary to meet the constraints of Excel sheet names,
        # or create a default one if none is supplied.
        #
        if sheet_name:
            sheet_name = mangle_sheet_name(sheet_name)
        else:
            sheet_name = u"Sheet %d" % n_sheet
        print("Sheet name: %s" % sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        yield "%s... " % sheet_name

        #
        # Write a single row containing the headers
        # Make the headers bold and freeze panes below that row
        #
        for i, (name, type) in enumerate(headers, 1):
            cell = ws.cell(column=i, row=1)
            cell.value = name
            cell.font = BOLD
            cell.fill = YELLOW
        ws.freeze_panes = "A2"

        yield "%s headers..." % sheet_name

        #
        # Append each row to the bottom of the sheet
        #
        n_row = 0
        for n_row, row in enumerate(rowset):
            if n_row > 0 and (n_row % ROWSET_SIZE == 0):
                yield "%s row %d" % (sheet_name, n_row)
            ws.append(munged(row))

        yield "%s %d rows" % (sheet_name, n_row)

        #
        # Get the max characters in each column. Then set the width to
        # that. Width is exactly the width of a monospace font (if not
        # changing other styles at least). Even if you use a variable width
        # font it is a decent estimation. This will not work with formulas.
        #
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        ws.auto_filter.ref = ws.dimensions

    yield "Save to %s" % spreadsheet_filepath
    wb.save(filename=spreadsheet_filepath)
